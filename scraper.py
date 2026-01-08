import pandas as pd
from bs4 import BeautifulSoup
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# ---------------- CONFIGURATION ----------------
SENDER_EMAIL = os.environ.get('EMAIL_USER')
SENDER_PASSWORD = os.environ.get('EMAIL_PASS')
RECEIVER_EMAIL = "priyansh.khandelwal@rosierfoods.com"
FILE_NAME = "Rosier_Report.xlsx"
# -----------------------------------------------

# 1. HTML File Load
try:
    with open("Amazon.html", "r", encoding="utf-8") as file:
        html_content = file.read()
except FileNotFoundError:
    print("Error: 'Amazon.html' repo me nahi mili.")
    exit()

soup = BeautifulSoup(html_content, 'html.parser')
products_data = []

product_divs = soup.find_all('div', {'data-component-type': 's-search-result'})
print(f"Products Found: {len(product_divs)}")

for item in product_divs:
    # Filter Logic (ROSIER)
    brand_tag = item.find('span', class_='a-size-base-plus a-color-base')
    if brand_tag and "ROSIER" in brand_tag.get_text().strip():
        
        # Name
        name_text = "Unknown"
        h2_tag = item.find('h2', class_='a-text-normal')
        if h2_tag and h2_tag.span:
            name_text = h2_tag.span.get_text().strip()

        # --- LINK EXTRACTION (FIXED HERE) ---
        product_link = None
        if h2_tag:
            parent_link = h2_tag.find_parent('a')
            if parent_link:
                # GALTI YAHAN THI: Sirf amazon.in jodna hai, search query nahi
                product_link = "https://www.amazon.in" + parent_link['href']

        # MRP
        mrp_text = "N/A"
        price_tag = item.find('span', class_='a-price-whole')
        if price_tag:
            mrp_text = price_tag.get_text().strip()

        # Stock
        stock_text = "Available"
        if item.find('span', class_='a-color-success'):
            stock_text = item.find('span', class_='a-color-success').get_text().strip()
        elif "Currently unavailable" in item.get_text():
            stock_text = "Out of Stock"

        products_data.append({
            'Brand': 'ROSIER',
            'Product Name': name_text,
            'MRP': mrp_text,
            'Stock': stock_text,
            'Hidden_URL': product_link
        })

# 2. Excel Creation with Working Hyperlinks
if products_data:
    df = pd.DataFrame(products_data)
    df.to_excel(FILE_NAME, index=False)
    
    # Fix Links using OpenPyXL
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    
    # Row 2 se start karenge
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # --- CONFIG: Kahan Link Lagana Hai? ---
        # Column B (Product Name) = row[1]
        # Column C (Price/MRP)    = row[2]
        
        target_cell = row[1]  # <--- Maine Product Name par link lagaya hai (Standard practice)
        # Agar Price par lagana hai to upar wali line ko 'row[2]' kar dein
        
        url_cell = row[4]     # Column E (Hidden_URL)
        
        if url_cell.value:
            target_cell.hyperlink = url_cell.value
            target_cell.font = Font(color="0000FF", underline="single")
            
    ws.delete_cols(5) # Remove URL column
    wb.save(FILE_NAME)
else:
    print("No ROSIER products found in the HTML file.")
    exit()

# 3. Email Sending
if not SENDER_EMAIL or not SENDER_PASSWORD:
    print("Error: GitHub Secrets set nahi hain.")
    exit()

msg = MIMEMultipart()
msg['From'] = SENDER_EMAIL
msg['To'] = RECEIVER_EMAIL
msg['Subject'] = "Daily Report: Amazon Scrap Data (Fixed Links)"
msg.attach(MIMEText("Hi Automailer,\n\nPFA Amazon Rosier products.", 'plain'))

with open(FILE_NAME, "rb") as attachment:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f"attachment; filename= {FILE_NAME}")
    msg.attach(part)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(SENDER_EMAIL, SENDER_PASSWORD)
server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())
server.quit()
print("Email Sent Successfully!")
